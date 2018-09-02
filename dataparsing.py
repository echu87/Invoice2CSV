from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os


wb = load_workbook(filename = "yes.xlsx")
sheet_ranges = wb['Sheet1']
print(sheet_ranges['A'].value)
